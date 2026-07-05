from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
INPUT = ROOT / "WBS_V3.xlsx"
OUTPUT = ROOT / "apps" / "api" / "prisma" / "wbs-import.json"


def rows(ws):
    header = None
    for row in ws.iter_rows(values_only=True):
        values = ["" if value is None else str(value).strip() for value in row]
        if header is None:
            header = values
            continue
        if not any(values):
            continue
        yield dict(zip(header, values))


def main():
    wb = load_workbook(INPUT, data_only=True)

    locations = []
    if "Lokasyon" in wb.sheetnames:
        for value in wb["Lokasyon"].iter_rows(values_only=True):
            code = value[0] if value else None
            if code:
                locations.append({"code": str(code).strip(), "name": f"Location {str(code).strip()}"})

    worker_types = []
    if "worker" in wb.sheetnames:
        for value in wb["worker"].iter_rows(values_only=True):
            name = value[0] if value else None
            if name:
                worker_types.append({"name": str(name).strip()})

    zzz_details = []
    if "Detay" in wb.sheetnames:
        seen = set()
        for value in wb["Detay"].iter_rows(values_only=True):
            code = value[0] if value else None
            if code:
                code = str(code).strip()
                if code not in seen:
                    seen.add(code)
                    zzz_details.append({"code": code, "name": f"ZZZ {code}"})

    type_of_works = []
    sub_type_of_works = []
    sub_sub_type_of_works = []
    if "Sub_Type" in wb.sheetnames:
        seen_tow = set()
        seen_stow = set()
        seen_sstow = set()
        for row_index, row in enumerate(wb["Sub_Type"].iter_rows(values_only=True), start=1):
            if row_index == 1:
                continue
            tow = row[0] if len(row) > 0 else None
            stow = row[1] if len(row) > 1 else None
            sstow = row[2] if len(row) > 2 else None
            unit = row[3] if len(row) > 3 else None
            type_code = row[4] if len(row) > 4 else None
            zzz = row[9] if len(row) > 9 else None

            if tow and tow not in seen_tow:
                seen_tow.add(tow)
                type_of_works.append({"code": str(tow).strip(), "name": str(tow).strip(), "sortOrder": len(type_of_works) + 1})
            if stow and stow not in seen_stow:
                seen_stow.add(stow)
                sub_type_of_works.append({
                    "code": str(stow).strip(),
                    "name": str(stow).strip(),
                    "typeOfWorkCode": str(tow).strip() if tow else "",
                    "sortOrder": len(sub_type_of_works) + 1,
                })
            if sstow and sstow not in seen_sstow:
                seen_sstow.add(sstow)
                sub_sub_type_of_works.append({
                    "code": str(sstow).strip(),
                    "name": str(sstow).strip(),
                    "subTypeOfWorkCode": str(stow).strip() if stow else "",
                    "unit": str(unit).strip() if unit else "",
                    "typeCode": str(type_code).strip() if type_code else "",
                    "zzzCode": str(zzz).strip() if zzz else None,
                })

    demo_users = [
        {"fullName": "Technical Office", "email": "techoffice@icn.com", "role": "TECH_OFFICE", "locationCodes": [loc["code"] for loc in locations[:2]]},
        {"fullName": "Head Of Master", "email": "hom@icn.com", "role": "HEAD_OF_MASTER", "locationCodes": [loc["code"] for loc in locations[:1]]},
        {"fullName": "Site Chief", "email": "sitechief@icn.com", "role": "SITE_CHIEF", "locationCodes": [loc["code"] for loc in locations[:2]]},
        {"fullName": "Project Manager", "email": "pm@icn.com", "role": "PROJECT_MANAGER", "locationCodes": [loc["code"] for loc in locations[:1]]},
        {"fullName": "Admin", "email": "admin@icn.com", "role": "ADMIN", "locationCodes": [loc["code"] for loc in locations[:1]]},
    ]

    payload = {
        "source": str(INPUT.name),
        "locations": locations,
        "projects": [{"code": f"PRJ-{loc['code']}", "name": f"Project {loc['code']}", "locationCode": loc["code"]} for loc in locations],
        "typeOfWorks": type_of_works,
        "subTypeOfWorks": sub_type_of_works,
        "subSubTypeOfWorks": sub_sub_type_of_works,
        "workerTypes": worker_types,
        "zzzDetails": zzz_details,
        "demoUsers": demo_users,
    }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()
